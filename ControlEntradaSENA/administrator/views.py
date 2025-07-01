from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth import authenticate, login, logout
from django.contrib import messages
import pandas as pd
from .models import *
from django.db import connection
from .forms import *
from django.db.models import Q
import openpyxl
from django.http import HttpResponse
from openpyxl.utils import get_column_letter
from datetime import date, timedelta
from django.db.models import Count
#Login admin
@user_passes_test(lambda u: not u.is_authenticated, login_url='adminpanel') #Si el usuario ya esta logeado no va a poder acceder a la pagina de login 
def login_admin(request):

    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        admin = authenticate(request, username=username, password=password)
        
        if admin is not None:
            login(request, admin)
            return redirect('adminpanel')
        else:
            messages.error(request, "Usuario o contraseña incorrectos")

    return render(request, 'login.html', {
        'title': 'Iniciar Sesion',
    })

#Logout admin
def logout_admin(request):
    logout(request)
    return redirect('login')

#Registro admin
def register_admin(request):
    form = RegisterForm(request.POST)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha registrado correctamente")
            return redirect('login')
        
    return render(request, 'register.html', {
        'title': 'Añadir administrador',
        'form': form
    })


#AdminPanel
@login_required(login_url="login")
def adminpanel(request):
    return render(request, 'adminpanel.html')

######

#Lista usuarios
@login_required(login_url="login")
def users(request):
    users = Usuarios.objects.all().prefetch_related('dispositivos_set').prefetch_related('vehiculos_set')
    
    search_query = request.GET.get('search')
    #Buscar usuarios
    if search_query:
        users = users.filter(
            Q(nombres__icontains=search_query) | Q(documento__icontains=search_query)
        )  
    if not users.exists():  #Luego revisamos si el resultado está vacío
        messages.warning(request, "No se encontraron usuarios con esa búsqueda.")
        
    checks = request.POST.get('checks-users')
    if checks:
        checks = checks.split(',')
        Usuarios.objects.filter(idusuario__in=checks).delete()
        messages.success(request, "Se han eliminado los usuarios seleccionados")

    delete = request.POST.get('delete-user')
    if delete:
        Usuarios.objects.filter(idusuario=delete).delete()
        messages.success(request, "Se ha eliminado el usuario")
    
    return render(request, 'pages/usuarios/users.html', {
        'title': 'Usuarios',
        'users': users
    })

#Registrar usuario
def register_user(request, rol):
    initial_data = {'rol': rol, 'rol_hide': rol}

    # ✅ Inyectar imagen correctamente
    if request.method == 'POST':
        request.FILES._mutable = True
        if 'foto_usuario' in request.FILES:
            request.FILES['imagen'] = request.FILES['foto_usuario']

    # ✅ Siempre crear el form con usuario_actual
    form = RegisterUser(
        request.POST or None,
        request.FILES or None,
        initial=initial_data,
        usuario_actual=request.user  # 🔥 clave
    )

    form.fields['centro'].required = rol != "3"
    form.fields['ficha'].required = rol == "2"

    if request.method == 'POST' and form.is_valid():
        user = form.save()
        messages.info(request, "Usuario registrado correctamente")
        return redirect('users')

    return render(request, 'pages/usuarios/register.html', {
        'title': 'Registrar usuario',
        'form': form,
        'rol': rol,
    })

#Editar usuario
def edit_user(request, id):
    instance = Usuarios.objects.get(idusuario=id)
    initial = {
        'imagen': instance.imagen,
        'rol': instance.rol,
        'rol_hide': instance.rol
    }
    
    form = RegisterUser(request.POST or None, request.FILES or None, instance=instance, initial=initial)

    form.fields['imagen'].required = False
    form.fields['centro'].required = False if instance.rol == 3 else True
    form.fields['ficha'].required = False if instance.rol != 2 else True
    form.fields['rol'].disabled = True  
    
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha editado correctamente")
            return redirect('users')

    return render(request, 'pages/usuarios/edit.html', {
        'title': 'Editar usuario',
        'form': form,
    })

###

# Lista de dispositivos
@login_required(login_url="login")
def dispositivo(request):
    devices = Dispositivos.objects.all() 
    query = request.GET.get('search', '')
    if query:
        devices = Dispositivos.objects.filter(sn__icontains=query)
        if not devices.exists():  #Luego revisamos si el resultado está vacío
            messages.warning(request, "No se encontraron dispositivos con esa búsqueda.")

    return render(request, 'pages/dispositivos/devices.html', {
        'title': 'Dispositivos',
        'dispositivos': devices,
        'search': query,
    })

def create_dispositivo(request):
    form = RegisterDevice(request.POST or None, request.FILES or None)
    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha registrado correctamente")
            return redirect('devices')
    return render(request, "pages/dispositivos/crear.html",{
        "form":form
    })
#Editar dispositivo
def edit_dispositivo(request, id):
    instance = Dispositivos.objects.get(iddispositivo=id)
    form = RegisterDevice(request.POST or None, request.FILES or None, instance= instance)
    if request .method == "POST" :


        if form.is_valid():
            form.save()
            messages.success(request, "se ha editado correctamente")
            return redirect('devices')
    return render(request, 'pages/dispositivos/edit.html', {
        'title': 'Editar dispositivo',
        'form': form
    })
#Eliminar dispositivo Individual
def delete_dispotivos(request, id):
        device = get_object_or_404(Dispositivos, iddispositivo=id)
        device.delete()
        return redirect("devices")

###

#Sanciones
@login_required(login_url="login")
def sanciones(request):
    sanciones = Sanciones.objects.all()
    
    return render(request, 'pages/sanciones/sanciones.html', {
        'title': 'Sanciones',
        'penaltys': sanciones
    })

# Crear Fichas
@login_required(login_url="login")
def crear_fichas(request):
    if request.method == 'POST':
        post_data = request.POST.copy()  # hacemos copia para modificarlo
        nombre_id = post_data.get('nombre')

        # Si el nombre no es un ID numérico, lo escribió el usuario y hay que crearlo
        if nombre_id and not nombre_id.isdigit():
            nuevo_nombre = FichasNombre.objects.create(nombre=nombre_id)
            post_data['nombre'] = nuevo_nombre.idfichanombre  # reescribimos con el ID real

        form = RegisterFicha(post_data)
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha creado correctamente")
            return redirect('fichas')
    else:
        form = RegisterFicha()

    return render(request, 'pages/fichas/crear.html', {
        'form': form
    })

#Fichas
@login_required(login_url="login")
def fichas(request):
    search_query = request.GET.get('search', '')
    fichas = Fichas.objects.all()
    form = CargarUsers()

    if search_query:
        fichas = fichas.filter(
            Q(nombre__nombre__icontains=search_query) |
            Q(jornada__nombre__icontains=search_query) |
            Q(numero__icontains=search_query) 
        )
        if not fichas.exists():
            messages.warning(request, "No se encontraron fichas con esa búsqueda.")

    return render(request, 'pages/fichas/fichas.html', {
        'title': 'Fichas',
        'fichas': fichas,
        'form': form    
    })


#Editar Fichas

def edit_ficha(request, id):
    instance = Fichas.objects.get(idficha=id)
    form = RegisterFicha(request.POST or None, instance=instance)

    if request.method == "POST":
        if form.is_valid():
            # Restauramos los campos deshabilitados manualmente
            form.instance.nombre = instance.nombre
            form.instance.tipo = instance.tipo
            form.save()
            messages.success(request, "Se ha editado correctamente")
            return redirect('fichas')

    return render(request, 'pages/fichas/edit.html', {
        'title': 'Editar Fichas',
        'form': form
    })

def cargar_users(request, ficha_id):
    ficha = get_object_or_404(Fichas, idficha=ficha_id)
    fichas = Fichas.objects.all()
    form = CargarUsers()

    if request.method == "POST":
        archivo_excel = request.FILES.get("archivo_excel")

        if not archivo_excel:
            return render(request, 'pages/fichas/fichas.html', {
                'error': "Debes subir un archivo Excel.",
                'form': form,
                'ficha': ficha,
                'fichas': fichas
            })

        try:
            df = pd.read_excel(archivo_excel, skiprows=9, engine='openpyxl')
            expected_columns = ['Tipo de Documento', 'Número de Documento', 'Nombre', 'Apellidos', 'Estado']

            if not all(col in df.columns for col in expected_columns):
                return render(request, 'pages/fichas/fichas.html', {
                    'error': "El archivo Excel no tiene las columnas esperadas.",
                    'ficha': ficha,
                    'fichas': fichas,
                    'form': form
                })

            rol = Roles.objects.get(idrol=2)  # Rol de aprendiz

            for _, row in df.iterrows():
                tipo_doc_instance = DocumentoTipo.objects.filter(nombre=row['Tipo de Documento']).first()
                if not tipo_doc_instance:
                    return render(request, 'pages/fichas/fichas.html', {
                        'error': f"Tipo de documento '{row['Tipo de Documento']}' no encontrado.",
                        'ficha': ficha,
                        'fichas': fichas,
                        'form': form
                    })

                estado = 1 if str(row['Estado']).strip().upper() == "EN FORMACION" else 0

                user, _ = Usuarios.objects.update_or_create(
                    documento=str(row['Número de Documento']),
                    defaults={
                        'nombres': row['Nombre'],
                        'apellidos': row['Apellidos'],
                        'tipodocumento': tipo_doc_instance,
                        'rol': rol,
                        'ficha': ficha,
                        'correo': None,
                        'telefono': None,
                        'centro': ficha.centro,
                        'imagen': 'images/users/default.jpg',  # Imagen genérica
                    }
                )

                use = FichasXAprendiz.objects.get_or_create(ficha=ficha, aprendiz=user)
                
            # Agregar el mensaje de éxito
            messages.success(request, 'Ficha cargada correctamente.')

        except Exception as e:
            return render(request, 'pages/fichas/fichas.html', {
                'error': f"Error al leer el archivo Excel: {str(e)}",
                'ficha': ficha,
                'fichas': fichas,
                'form': form
            })

        return redirect('fichas')

    return render(request, 'pages/fichas/fichas.html', {
        'ficha': ficha,
        'fichas': fichas,
        'form': form
    })

#Editar sanciones

def edit_sanciones(request, id):
    instance=Sanciones.objects.get(idsancion=id)

    form = RegisterSanciones(request.POST or None, request.FILES or None, instance= instance)
    if request .method == "POST" :
        if form.is_valid():
            form.save()
            messages.success(request, "se ha editado correctamente")
            return redirect('sanciones')
            
    return render(request, 'pages/sanciones/edit.html', {
        'title': 'Editar Sanciones',
        'form': form
     })


#Vehicles
@login_required(login_url="login")
def vehicles(request):
    search_query = request.GET.get('search', '')  # Obtiene el término de búsqueda desde la URL
    print(f"Search query: {search_query}")  # Para verificar que estamos recibiendo la búsqueda
    
    # Inicializar vehicles por defecto
    vehicles = Vehiculos.objects.all()  # Por defecto, mostrar todos los vehículos
    
    if search_query:  # Si hay un término de búsqueda
        vehicles = Vehiculos.objects.filter(placa__icontains=search_query)  # Filtra por placa
    if not vehicles:
        messages.warning(request, "No se encontraron vehiculos con esa búsqueda.")
    
    return render(request, 'pages/vehiculos/vehicles.html', {
        'title': 'Vehículos',
        'vehicles': vehicles
    })
def create_vehicle(request):
    rol = None

    # Si es superuser (admin)
    if request.user.is_authenticated and request.user.is_superuser:
        rol = 'admin'

    form = RegisterVehicle(request.POST or None, request.FILES or None, rol=rol)

    if request.method == 'POST':
        if form.is_valid():
            form.save()
            messages.success(request, "Se ha registrado correctamente")
            return redirect('vehiculos')

    return render(request, "pages/vehiculos/crear.html", {
        "form": form
    })
#Editar vehiculos
def edit_vehiculo(request, id):
    instance=Vehiculos.objects.get(idvehiculo=id)

    form = RegisterVehicle(request.POST or None, request.FILES or None, instance= instance)
    if request .method == "POST" :
        if form.is_valid():
            form.save()
            messages.success(request, "Vehículo editado correctamente")
            return redirect('vehiculos')
            
    return render(request, 'pages/vehiculos/edit.html', {
         'title': 'Editar Vehiculos',
         'form': form 
    })

#Eliminar vehiculos Individual
def delete_vehiculo(request, id):
    device = get_object_or_404(Vehiculos, idvehiculo =id)
    device.delete()
    return redirect("vehiculos")



#acerda de
@login_required(login_url="login")
def about(request):

    return render(request, 'pages/acerca/about.html', {
        'title': 'Acerca de'
    })


#Reportes
@login_required(login_url="login")
def reportes(request):
    search_query = request.GET.get('search', '')  # Obtiene el término de búsqueda desde la URL
    print(f"Search query: {search_query}")  # Para verificar que estamos recibiendo la búsqueda
    
    # Inicializar vehicles por defecto
    salidas = Salidas.objects.all() # Por defecto, mostrar todos los vehículos
    
    if search_query:  # Si hay un término de búsqueda
       salidas = Salidas.objects.filter(ingreso__usuario__documento__icontains=search_query)
 # Filtra por documento de usuario
    if not salidas.exists():
        messages.warning(request, "No se encontraron salidas con esa búsqueda.")
    
    
    return render(request, 'pages/reportes/reportes.html', {
        'title': 'Reportes',
        'salidas':salidas,

    })
def reporteInstructor(request):


    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Roles Administrativos"

    # Escribir encabezados
    encabezados = [
        # ingresos
        'Número Documento', 'Nombre',  'Fecha Ingreso', 'Hora Entrada',
        'Dispositivo 1 SN', 'Dispositivo 2 SN', 'Dispositivo 3 SN', 'Vehiculo',
        # salidas
        'Fecha Salida', 'Hora Salida', 'Dispositivo 1 Salida SN', 'Dispositivo 2 Salida SN', 'Dispositivo 3 Salida SN'
    ]
    ws.append(encabezados)

    # Escribir datos
    salidas = Salidas.objects.filter(
    Q(ingreso__usuario__rol=1) | Q(ingreso__usuario__rol=4)
    )
    for salida in salidas:  
        #vehiculo 
        vehiculo = salida.ingreso.vehiculo.placa if salida.ingreso.vehiculo else "None"
        # Dispositivos de ingreso
        dispositivo1i = salida.ingreso.dispositivo.sn if salida.ingreso.dispositivo else "None"
        dispositivo2i = salida.ingreso.dispositivo2.sn if salida.ingreso.dispositivo2 else "None"
        dispositivo3i = salida.ingreso.dispositivo3.sn if salida.ingreso.dispositivo3 else "None"
        # Dispositivos de salida
        dispositivo1s = salida.dispositivo.sn if salida.dispositivo else "None"
        dispositivo2s = salida.dispositivo2.sn if salida.dispositivo2 else "None"
        dispositivo3s = salida.dispositivo3.sn if salida.dispositivo3 else "None"
    
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.append([
            # Ingresos
            salida.ingreso.usuario.documento,
            salida.ingreso.usuario.nombres,
            salida.ingreso.fecha,
            salida.ingreso.horaingreso,
            dispositivo1i,
            dispositivo2i,
            dispositivo3i,
            vehiculo,
            # salidas
            salida.fecha,
            salida.horasalida,
            dispositivo1s,
            dispositivo2s,
            dispositivo3s


        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    fecha_actual = date.today().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Admins_{fecha_actual}.xlsx"'

    wb.save(response)

    return response

def reporteAprendiz(request):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aprendices"
    # Escribir encabezados
    encabezados = [
        # ingresos
        'Número Documento', 'Nombre', 'Ficha', 'Fecha Ingreso', 'Hora Entrada',
        'Dispositivo 1 SN', 'Dispositivo 2 SN', 'Dispositivo 3 SN', 'Vehiculo',
        # salidas
        'Fecha Salida', 'Hora Salida', 'Dispositivo 1 Salida SN', 'Dispositivo 2 Salida SN', 'Dispositivo 3 Salida SN'
    ]
    ws.append(encabezados)
    # Escribir datos
    salidas = Salidas.objects.filter(
    Q(ingreso__usuario__rol=2) | Q(ingreso__usuario__rol=3)
    )
    for salida in salidas:  
        #vehiculo 
        vehiculo = salida.ingreso.vehiculo.placa if salida.ingreso.vehiculo else "None"
        # Dispositivos de ingreso
        dispositivo1i = salida.ingreso.dispositivo.sn if salida.ingreso.dispositivo else "None"
        dispositivo2i = salida.ingreso.dispositivo2.sn if salida.ingreso.dispositivo2 else "None"
        dispositivo3i = salida.ingreso.dispositivo3.sn if salida.ingreso.dispositivo3 else "None"
        # Dispositivos de salida
        dispositivo1s = salida.dispositivo.sn if salida.dispositivo else "None"
        dispositivo2s = salida.dispositivo2.sn if salida.dispositivo2 else "None"
        dispositivo3s = salida.dispositivo3.sn if salida.dispositivo3 else "None"
        # Ficha
        ficha = salida.ingreso.usuario.ficha.numero if salida.ingreso.usuario.ficha else "None"
        for col in range(1, len(encabezados) + 1):
            ws.column_dimensions[get_column_letter(col)].width = 20
        ws.append([
            # Ingresos
            salida.ingreso.usuario.documento,
            salida.ingreso.usuario.nombres,
            ficha,
            salida.ingreso.fecha,
            salida.ingreso.horaingreso,
            dispositivo1i,
            dispositivo2i,
            dispositivo3i,
            vehiculo,
            # salidas
            salida.fecha,
            salida.horasalida,
            dispositivo1s,
            dispositivo2s,
            dispositivo3s
        ])
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    fecha_actual = date.today().strftime('%Y-%m-%d')
    response['Content-Disposition'] = f'attachment; filename="Reporte_Aprendices_{fecha_actual}.xlsx"'
    wb.save(response)

    return response



# Informe estadistico
def daterange(start_date, end_date):
    for n in range(int((end_date - start_date).days) + 1):
        yield start_date + timedelta(n)

def informe_estadistico(request):
    hoy = date.today()
    hace_7_dias = hoy - timedelta(days=6)

    # Crear lista con los últimos 7 días como strings (para labels)
    fechas = [d.strftime('%Y-%m-%d') for d in daterange(hace_7_dias, hoy)]

    # Ingresos por fecha
    ingresos_query = (
        Ingresos.objects
        .filter(fecha__range=[hace_7_dias, hoy])
        .values('fecha')
        .annotate(total=Count('idingreso'))
    )
    ingresos_dict = {item['fecha'].strftime('%Y-%m-%d'): item['total'] for item in ingresos_query}

    ingresos_dia = [ingresos_dict.get(fecha, 0) for fecha in fechas]

    # Salidas por fecha
    salidas_query = (
        Salidas.objects
        .filter(fecha__range=[hace_7_dias, hoy])
        .values('fecha')
        .annotate(total=Count('idsalida'))
    )
    salidas_dict = {item['fecha'].strftime('%Y-%m-%d'): item['total'] for item in salidas_query}
    salidas_dia = [salidas_dict.get(fecha, 0) for fecha in fechas]

    # Ingresos por rol de usuario (últimos 7 días)
    ingresos_rol_qs = (
        Ingresos.objects
        .filter(fecha__range=[hace_7_dias, hoy])
        .values('usuario__rol')
        .annotate(total=Count('idingreso'))
    )
    ingresos_rol = {item['usuario__rol']: item['total'] for item in ingresos_rol_qs}

    # Salidas por rol de usuario (últimos 7 días)
    salidas_rol_qs = (
        Salidas.objects
        .filter(fecha__range=[hace_7_dias, hoy])
        .values('ingreso__usuario__rol')
        .annotate(total=Count('idsalida'))
    )
    salidas_rol = {item['ingreso__usuario__rol']: item['total'] for item in salidas_rol_qs}

    # Totales generales
    total_ingresos = Ingresos.objects.count()
    total_salidas = Salidas.objects.count()

    context = {
        'fechas': fechas,
        'ingresos_dia': ingresos_dia,
        'salidas_dia': salidas_dia,
        'ingresos_rol': ingresos_rol,
        'salidas_rol': salidas_rol,
        'total_ingresos': total_ingresos,
        'total_salidas': total_salidas,
    }

    return render(request, 'pages/reportes/estadistico.html', context)